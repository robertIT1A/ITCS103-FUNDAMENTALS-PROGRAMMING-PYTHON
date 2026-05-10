import openpyxl as op


wk = op.Workbook()
sheet = wk.active

sheet ['A1'] = "ID"
sheet ['B1'] = "Last Name"
sheet['C1'] = "First Name"
sheet['D1'] = "Birth Year"
sheet['E1'] = "Age"


for i in range(1,3+1):
    print(f"Person {i}")
    one_fname = input("Enter First name: ")
    one_lname = input("Enter Last name: ")
    one_birth = int(input("Enter birth year: "))

    new_id = sheet.max_row
    age1 = 2026 - one_birth
    sheet.append([new_id, one_lname, one_fname,one_birth,age1])
    wk.save("favorite_people.xlsx")
    print("\n")


print("\nFavorite people saved successfully!\n")
print("=== FAVORITE PEOPLE LIST ===\n")


workbook = op.load_workbook("favorite_people.xlsx")
sheet = workbook.active



for row in sheet.iter_rows(values_only=True):
    print(row)

alis = input("\nPress Enter to exit....")


